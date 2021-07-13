from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook


class ExcelImporter(object):
    """
    认为导入文件每行应该是一个对象
    最后返回以key为excel中行数的字典

    imported_fields如果提供，则可以根据提供的field返回

    举例excel为
        ----------------------------------------------------
        活动名称    活动描述        开始时间        结束时间
        ----------------------------------------------------
        兰蔻满减    这是一个活动     2020-10-11     2050-10-11
        兰蔻满减2   这是一个活动     2020-10-11     2050-10-11


    imported_fields 提供，内容是
        {
            'title': '活动名称',
            'desc': '活动描述',
            'start_time': '开始时间',
            'end_time': '结束时间'
        }

    其中value是被认为在excel第一行中的文字内容，key在返回数据中仍然用作key

    返回结果为

        {
            '2': {
                'title': '兰蔻满减',
                'desc': '这是一个活动',
                'start_time': '2020-10-11',
                'end_time': '2050-10-11'
            },
            '3': {
                'title': '兰蔻满减2',
                'desc': '这是一个活动',
                'start_time': '2020-10-11',
                'end_time': '2050-10-11'
            }
        }

    如果不提供imported_fields，则返回每行每列：
        {
            '1': {
                'A': 'title',
                'B': '名字',
                'C': '描述'
            },
            '2': {
                'A': 'title2',
                'B': '名字2',
                'C': '描述2'
            }
        }

    Note:
    认为excel第一行必须是column的title，且需要和提供的imported_fields一致
    如果不关心title，第一行也不能放数据，数据只从第二行开始读

    如果传入了imported_fields，实际excel多余的column（imported_fields里没有的项）不会被读或记录，
    返回内容会以imported_fields里的field为准

    """
    def __init__(self, imported_fields={}):
        super(ExcelImporter, self).__init__()
        self.imported_fields = imported_fields

    def import_file(self, file):
        '''
        只解析文件，解析后返回一个字典，每一行应该是一个对象
        '''
        result = OrderedDict()
        # 记录所在列 fields_columns = {'title':'D', 'name':'E', 'desc':'F'}
        fields_columns = dict()
        wb = load_workbook(file, read_only=False, data_only=True)
        # 模板里只用一页
        ws = wb.worksheets[0]
        # 把每列对应什么字段记录在fields_columns
        for column in ws.iter_cols():
            for field_key, field_value in self.imported_fields.items():
                # 认为只有第一行是title
                if column[0].value == field_value.strip():
                    fields_columns.update({
                        field_key: column[0].column # 'title': 'D'
                    })
                    break
        # 按照每行读入
        # 如果有指定的fields
        if self.imported_fields:
            for row_index in range(2, ws.max_row + 1):
                fields_row = OrderedDict()
                for field_name, column_index in fields_columns.items():
                    value = ws['%s%s' % (column_index, row_index)].value
                    if value is not None:
                        field = {field_name: ws['%s%s' % (column_index, row_index)].value}
                        fields_row.update(field)
                result.update({
                    str(row_index): fields_row
                })
        # 如果没有指定fields则全返回
        else:
            for row_index in range(ws.min_row, ws.max_row + 1):
                fields_row = OrderedDict()
                for column_index in range(ws.min_column - 1, ws.max_column):
                    cell = ws[row_index][column_index]
                    field = {cell.column: cell.value}
                    fields_row.update(field)
                result.update({
                    str(row_index): fields_row
                })
        return result
    
    def import_file_to_list(self, file):
        '''
        只解析文件，解析后返回一个列表，按照顺序对应Excel每行数据
        '''
        result_new = []
        # 记录所在列 fields_columns = {'title':'D', 'name':'E', 'desc':'F'}
        fields_columns = dict()
        wb = load_workbook(file, read_only=False, data_only=True)
        # 模板里只用一页
        ws = wb.worksheets[0]
        #import pdb; pdb.set_trace()
        # 把每列对应什么字段记录在fields_columns
        for column in ws.iter_cols():
            for field_key, field_value in self.imported_fields.items():
                # 认为只有第一行是title
                if column[0].value == field_value.strip():
                    fields_columns.update({
                        field_key: column[0].column # 'title': 'D'
                    })
                    break
        # 按照每行读入
        # 如果有指定的fields
        if self.imported_fields:
            for row_index in range(2, ws.max_row + 1):
                fields_row = OrderedDict()
                for field_name, column_index in fields_columns.items():
                    value = ws['%s%s' % (column_index, row_index)].value
                    if value is not None:
                        field = {field_name: ws['%s%s' % (column_index, row_index)].value}
                        fields_row.update(field)
                result_new.append(fields_row)

        # 如果没有指定fields则全返回
        else:
            for row_index in range(ws.min_row, ws.max_row + 1):
                fields_row = OrderedDict()
                for column_index in range(ws.min_column - 1, ws.max_column):
                    cell = ws[row_index][column_index]
                    field = {cell.column: cell.value}
                    fields_row.update(field)
                result_new.append(fields_row)

        return result_new

    def import_multi_title_file(self, file, title_row):
        '''
        只解析文件，解析后返回一个字典，每一行应该是一个对象
        '''
        # import pdb; pdb.set_trace()
        if not self.imported_fields:
            raise Exception("表格标题未设置")
        result = OrderedDict()
        # 记录所在列 fields_columns = {'title':'D', 'name':'E', 'desc':'F'}
        fields_columns = dict()
        wb = load_workbook(file, read_only=True, data_only=True)
        # 模板里只用一页
        ws = wb.worksheets[0]
        if ws.max_row < 3:
            raise Exception("表格行数太少，无法读取信息")
        # 把每列对应什么字段记录在fields_columns
        count = 0
        for row in ws.rows:
            count += 1
            if count == title_row:
                for field_key, field_value in self.imported_fields.items():

                    for cell in row:
                        # 认为只有第一行是title
                        if cell.value == field_value.strip():
                            fields_columns.update({
                                field_key: cell.column # 'title': 'D'
                            })
                            break
                break

        if not fields_columns:
            raise Exception("无法找到表格标题")
        # 按照每行读入
        count = 0
        for row in ws.rows:
            count += 1
            if count < title_row + 1:
                continue
            
            fields_row = OrderedDict()
            for field_name, column_index in fields_columns.items():
                
                value = ws[count][column_index-1].value
                if value is not None:
                    field = {field_name: value}
                    fields_row.update(field)
                result.update({
                    str(count): fields_row
                })
        return result

    def export_file(self):
        pass


# class ExcelExporter(object):
#     """
#     exported_fields 应该是一个字典，内容是属性或者变量所对应的字段
#     e.g.
#         exported_fields = {
#             'title': '商品名称',
#             'sku_attr': '商品属性'
#         }
#     之后
#     items 应该是 products
#     获取每行数据的时候，会迭代传入的items，用exported_fields里的key作为要获取的变量或属性

#     exported_fields 为空的话，items直接输出

#     2018-07-11
#     试验过，发现传进来 exported_fields 再变OrderedDict，是没办法和items顺序一致的
#     所以做的这么复杂不如直接用户拼好数据了

#     """
#     def __init__(self, items, exported_fields={}):
#         super(ExcelImporter, self).__init__()
#         self.exported_fields = OrderedDict(exported_fields)
#         self.items = items

#     def validate_data(self):
#         '''
#         验证 title 的数量 和 data里面一条的数量对的上
#         '''
#         if self.exported_fields:
#             first_data = self.items[0]
#             if len(first_data) != len(self.exported_fields.keys()):
#                 raise ValueError('传入exported_fields和数据items一行的数量不一致。')

#     def export_file(self):
#         '''
#         title 放置导出excel里第一行想展示的文字，可以为空
#         data 应该是一个list，可迭代，每一个元素就是一行
#         '''
#         self.validate_data()
#         wb = Workbook()
#         ws = wb.active
#         # result = []
#         if self.exported_fields:
#             titles = self.exported_fields.values()
#             import pdb; pdb.set_trace()
#             ws.append(titles)
#         for item in self.items:


def export_file(items):
    '''
    '''
    wb = Workbook()
    ws = wb.active
    for item in items:
        ws.append(item)
    result = save_virtual_workbook(wb)
    return result
    # response = HttpResponse(
    #     result, content_type='application/vnd.ms-excel', status=200)
    # return response


def export_self_service_file(items):
    wb = Workbook()
    ws = wb.active
    for item in items:
        ws.append(item)
    font = Font(bold=True, size=18)
    # 合并单元格，设置单元格值居中
    ws.merge_cells("A1:F2")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = font
    ws.merge_cells("A3:C4")
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D3:E4")
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("F3:F4")
    ws["F3"].alignment = Alignment(horizontal="center", vertical="center")
    last_row = len(items)
    ws.merge_cells("A{}:D{}".format(last_row-1, last_row))
    ws.merge_cells("E{}:F{}".format(last_row-1, last_row))
    # import pdb; pdb.set_trace()
    # 设置边框
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

    for row_count, row in enumerate(ws.rows):
        if row_count < 4:
            for cell in row:
                cell.border = border
        else:
            for count, cell in enumerate(row):
                if count == 0:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
    
    ws["A{}".format(last_row-1)].alignment = Alignment(horizontal="left", vertical="center")
    ws["A{}".format(last_row)].border = border 
    ws["B{}".format(last_row)].border = border 
    ws["C{}".format(last_row)].border = border 
    ws["D{}".format(last_row)].border = border 
    ws["E{}".format(last_row)].border = border 
    ws["F{}".format(last_row)].border = border 

    # 设置列宽
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 8.75
    ws.column_dimensions["C"].width = 8.75
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 31.25
    result = save_virtual_workbook(wb)
    return result


def export_sf_file(items):
    wb = Workbook()
    ws = wb.active
    for item in items:
        ws.append(item)
    # 合并单元格，设置单元格值居中
    ws.merge_cells("G1:J1")
    ws.merge_cells("K1:N1")
    ws.merge_cells("O1:T1")
    ws.merge_cells("U1:AZ1")
    result = save_virtual_workbook(wb)
    return result